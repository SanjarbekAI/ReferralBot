from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook


async def export_users_registered_bot(data):
    wb = Workbook()

    # # get active sheet
    ws = wb.active
    ws.title = "Users List"

    # extend columns size
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 25

    # creating list for excel file column headers
    headings = ["Ism Familiya", "Telefon raqam", "Vazn", "Ro'xatdan o'tgan vaqti"]
    ws.append(headings)

    # changing column header text to bold
    for col in range(1, 10):
        ws[get_column_letter(col) + '1'].font = Font(bold=True)

    for col in ws.iter_cols(min_col=1, max_col=len(headings)):
        for cell in col:
            cell.alignment = Alignment(horizontal='center')

    i = 2
    for user in data:
        print(user['weight'])
        current = ws[f"A{i}"]
        current.alignment = Alignment(horizontal='center', vertical='center')
        ws[f"A{i}"].value = user["full_name"]

        current = ws[f"B{i}"]
        current.alignment = Alignment(horizontal='center', vertical='center')
        ws[f"B{i}"].value = user["phone_number"]

        current = ws[f"C{i}"]
        current.alignment = Alignment(horizontal='center', vertical='center')
        ws[f"C{i}"].value = user["weight"]

        current = ws[f"D{i}"]
        current.alignment = Alignment(horizontal='center', vertical='center')
        ws[f"D{i}"].value = str(user["created_at"])[0:19]

        i += 1
    else:
        pass

    workbook = save_virtual_workbook(wb)
    return workbook
